from fastapi import APIRouter, File, UploadFile, Depends, HTTPException
from openpyxl import load_workbook
from ..database import SessionLocal
from .. import schemas, crud
from ..auth_utils import get_current_user
from ..audit import record_audit

router = APIRouter(prefix='/import', tags=['import'])

@router.post('/excel')
def import_excel(file: UploadFile = File(...), user = Depends(get_current_user)):
    if not file.filename.endswith(('.xlsx', '.xlsm')):
        raise HTTPException(status_code=400, detail='Only xlsx files supported')
    wb = load_workbook(filename=file.file, data_only=True)
    ws = wb.active
    created = 0
    db = SessionLocal()
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            title = row[0]
            description = row[1] if len(row) > 1 else None
            owner = row[2] if len(row) > 2 else None
            due_date = row[3] if len(row) > 3 else None
            if not title:
                continue
            req = schemas.AuditRequestCreate(title=str(title), description=str(description) if description else None, owner=str(owner) if owner else None, due_date=str(due_date) if due_date else None)
            crud.create_audit_request(db, req)
            created += 1
    finally:
        db.close()
    record_audit(actor=user.username, action='IMPORT', resource='Excel', summary=f'Imported {created} rows')
    return {'status':'ok','created':created}
